"""Export a stress-battery results JSON to an Excel file with one
row per question and ALL fields including the full answer text."""
import json, sys
from pathlib import Path
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

src = sys.argv[1] if len(sys.argv) > 1 else "/tmp/test_suite_200_results.json"
dst = sys.argv[2] if len(sys.argv) > 2 else "/tmp/MISA_Test_Suite_200_Results.xlsx"

records = json.load(open(src))
df = pd.DataFrame(records)
# Reorder columns so Question + Answer are first
ordered = [
    "original_question", "question", "label", "flags", "intent",
    "elapsed_s", "answer_chars", "answer", "answer_head",
    "was_substituted", "has_history", "status",
]
ordered = [c for c in ordered if c in df.columns]
df = df[ordered + [c for c in df.columns if c not in ordered]]
# Flags column to comma-separated string
if "flags" in df.columns:
    df["flags"] = df["flags"].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)

with pd.ExcelWriter(dst, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="200 Question Results", index=False)
    ws = writer.sheets["200 Question Results"]
    # Format header row
    header_fill = PatternFill("solid", start_color="1A365D")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Set widths
    widths = {
        "original_question": 55, "question": 55, "label": 10, "flags": 25,
        "intent": 18, "elapsed_s": 10, "answer_chars": 10,
        "answer": 80, "answer_head": 55,
        "was_substituted": 12, "has_history": 12, "status": 8,
    }
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 15)
    # Wrap answer columns + tall rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    # Freeze top row
    ws.freeze_panes = "A2"
    # Color-code label cells
    fills = {"good": "C6F6D5", "mediocre": "FEF3C7", "broken": "FED7D7"}
    label_col_idx = list(df.columns).index("label") + 1
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=label_col_idx).value
        if v in fills:
            ws.cell(row=r, column=label_col_idx).fill = PatternFill("solid", start_color=fills[v])

    # Summary sheet
    summary_df = pd.DataFrame([
        {"Metric": "Total questions", "Value": len(df)},
        {"Metric": "Good", "Value": int((df["label"] == "good").sum())},
        {"Metric": "Mediocre", "Value": int((df["label"] == "mediocre").sum())},
        {"Metric": "Broken", "Value": int((df["label"] == "broken").sum())},
        {"Metric": "Avg latency (s)", "Value": round(float(df["elapsed_s"].mean()), 1)},
        {"Metric": "p50 latency (s)", "Value": round(float(df["elapsed_s"].median()), 1)},
        {"Metric": "p95 latency (s)", "Value": round(float(df["elapsed_s"].quantile(0.95)), 1)},
        {"Metric": "Max latency (s)", "Value": round(float(df["elapsed_s"].max()), 1)},
    ])
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    s = writer.sheets["Summary"]
    s.column_dimensions["A"].width = 28
    s.column_dimensions["B"].width = 15
    for col_idx in (1, 2):
        cell = s.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

print(f"Saved: {dst}")
print(f"Total rows: {len(df)}")
print(f"  Good:     {(df['label'] == 'good').sum()}")
print(f"  Mediocre: {(df['label'] == 'mediocre').sum()}")
print(f"  Broken:   {(df['label'] == 'broken').sum()}")
